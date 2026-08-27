#!/usr/bin/env python3
"""
run_portfolio_optimizer.py  (ltbbot)

Laedt alle Envelope-Configs, fuehrt Portfolio-Simulation durch und waehlt
das beste Portfolio per Greedy-Algorithmus. Schreibt active_strategies in
settings.json.

Aufruf:
  python3 run_portfolio_optimizer.py              # interaktiv
  python3 run_portfolio_optimizer.py --auto-write # automatisch (Scheduler)
  python3 run_portfolio_optimizer.py --replot     # Replot aktives Portfolio
"""
import contextlib
import io
import os
import sys
import json
import argparse
from datetime import date, timedelta
from tqdm import tqdm

PROJECT_ROOT  = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))

CONFIGS_DIR   = os.path.join(PROJECT_ROOT, 'src', 'ltbbot', 'strategy', 'configs')
SETTINGS_PATH = os.path.join(PROJECT_ROOT, 'settings.json')

B  = '\033[1;37m'
G  = '\033[0;32m'
Y  = '\033[1;33m'
R  = '\033[0;31m'
NC = '\033[0m'

# Muss deckungsgleich mit dem lookback_days-Mapping in run_pipeline.sh sein (das dort
# jede Symbol/Timeframe-Optimierung inkl. deren IS/OOS-Split bestimmt) -- sonst waehlt/
# bestaetigt der Portfolio-Optimizer Strategien auf einem ANDEREN Trainingsfenster als
# dem, auf dem sie ueberhaupt gefittet und OOS-bestaetigt wurden. Vorher inkonsistent
# (u.a. 1h: 365 hier vs. 548 in run_pipeline.sh, 4h: 730 vs. 1095, 1d: 1095 vs. 1825) --
# angeglichen 2026-08-26 (PIPELINE_UPDATE_AND_28PAIR_SWEEP_2026-08.md).
LOOKBACK_MAP = {
    '5m': 90,   '15m': 90,
    '30m': 548, '1h': 548,
    '2h': 730,
    '4h': 1095, '6h': 1095,
    '1d': 1825,
}
BOT_NAME = 'ltbbot'


def _scan_configs() -> list:
    if not os.path.isdir(CONFIGS_DIR):
        return []
    return sorted([
        os.path.join(CONFIGS_DIR, f)
        for f in os.listdir(CONFIGS_DIR)
        if f.endswith('_envelope.json')
    ])


def _build_strategies_data(config_files: list, start_date: str, end_date: str) -> dict:
    from ltbbot.analysis.backtester import load_data, FINE_TF_MAP, LazyFineData
    strategies_data = {}
    for path in tqdm(config_files, desc='Lade Configs & Daten'):
        fname = os.path.basename(path)
        try:
            with open(path) as f:
                config = json.load(f)
            market    = config.get('market', {})
            symbol    = market.get('symbol', '')
            timeframe = market.get('timeframe', '')
            if not symbol or not timeframe:
                continue
            data = load_data(symbol, timeframe, start_date, end_date)
            if data is None or data.empty or len(data) < 50:
                print(f"  {Y}Uebersprungen (keine Daten): {fname}{NC}")
                continue

            fine_tf = FINE_TF_MAP.get(timeframe)
            fine_data = LazyFineData(symbol, fine_tf) if fine_tf else None

            # portfolio_simulator erwartet 'params' als vollstaendiges Config-Dict
            strategies_data[fname] = {
                'symbol':    symbol,
                'timeframe': timeframe,
                'data':      data,
                'fine_data': fine_data,
                'params':    config,
            }
        except Exception as e:
            print(f"  {Y}Fehler bei {fname}: {e}{NC}")
    return strategies_data


def _write_to_settings(portfolio_files: list, strategies_data: dict) -> None:
    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    existing     = settings.get('live_trading_settings', {}).get('active_strategies', [])
    existing_map = {(s.get('symbol'), s.get('timeframe')): s for s in existing}
    new_strategies = []
    for fname in portfolio_files:
        sd        = strategies_data.get(fname, {})
        symbol    = sd.get('symbol', '')
        timeframe = sd.get('timeframe', '')
        if not symbol or not timeframe:
            continue
        base  = existing_map.get((symbol, timeframe), {})
        entry = {**base, 'symbol': symbol, 'timeframe': timeframe, 'active': True}
        new_strategies.append(entry)
    lt = settings.setdefault('live_trading_settings', {})
    lt['active_strategies']          = new_strategies
    lt['use_auto_optimizer_results'] = True
    with open(SETTINGS_PATH, 'w') as f:
        json.dump(settings, f, indent=4)


def _simulate_current_portfolio(settings: dict, strategies_data: dict,
                                 start_capital: float,
                                 start_date: str, end_date: str):
    from ltbbot.analysis.portfolio_simulator import run_portfolio_simulation
    current = [
        s for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
        if s.get('active')
    ]
    if not current:
        return None
    sim_data = {}
    for s in current:
        sym, tf = s.get('symbol', ''), s.get('timeframe', '')
        for fname, sd in strategies_data.items():
            if sd['symbol'] == sym and sd['timeframe'] == tf:
                sim_data[fname] = sd
                break
    if not sim_data:
        return None
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        return run_portfolio_simulation(start_capital, sim_data, start_date, end_date,
                                         multi_band_entries=True)


def _get_telegram_creds():
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json')) as f:
            s = json.load(f)
        tg = s.get('telegram', {})
        t, c = tg.get('bot_token', ''), tg.get('chat_id', '')
        return (t, c) if t and c else (None, None)
    except Exception:
        return None, None


def _send_telegram(msg):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        requests.post(f'https://api.telegram.org/bot{token}/sendMessage',
                      data={'chat_id': chat, 'text': msg}, timeout=10)
    except Exception:
        pass


def _send_telegram_doc(fpath, caption=''):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        with open(fpath, 'rb') as fh:
            requests.post(f'https://api.telegram.org/bot{token}/sendDocument',
                          data={'chat_id': chat, 'caption': caption},
                          files={'document': fh}, timeout=30)
    except Exception:
        pass


def generate_trades_excel(final, strategies_data, capital, start_date, end_date):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f'  {Y}openpyxl nicht installiert — Excel uebersprungen.{NC}')
        return None

    # `trades_df` ist der tatsaechliche Schluessel aus run_portfolio_simulation()
    # (portfolio_simulator.py) -- eine DataFrame mit 'side'/'pnl_usd' statt
    # 'direction'/'pnl'. Der vorherige Code las 'trade_history'/'direction'/'pnl',
    # die es nie gab -- die Excel-Generierung war dadurch IMMER leer (auch im
    # --replot-Pfad, nicht nur --auto-write), unabhaengig vom heutigen Fix.
    trades_raw = final.get('trades_df')
    if trades_raw is None or (hasattr(trades_raw, 'empty') and trades_raw.empty):
        return None
    trades = trades_raw.sort_values('exit_time').to_dict('records') if hasattr(trades_raw, 'sort_values') else trades_raw
    if not trades:
        return None

    equity = capital
    rows = []
    for i, t in enumerate(trades, 1):
        pnl    = t.get('pnl_usd', t.get('pnl', 0.0))
        symbol = t.get('symbol', '?')
        equity += pnl
        rows.append({
            'Nr':            i,
            'Datum':         str(t.get('exit_time', t.get('entry_time', '')))[:16].replace('T', ' '),
            'Coin':          symbol.split('/')[0] if '/' in symbol else symbol,
            'Symbol':        symbol,
            'Timeframe':     t.get('timeframe', '?'),
            'Richtung':      str(t.get('side', t.get('direction', '?'))).upper(),
            'Ergebnis':      'TP erreicht' if str(t.get('reason', '')).upper() == 'WIN' else 'SL erreicht',
            'Entry-Preis':   t.get('entry_price', 0.0),
            'Exit-Preis':    t.get('exit_price', 0.0),
            'PnL (USDT)':    round(pnl, 4),
            'Gesamtkapital': round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'
    hdr  = PatternFill('solid', fgColor='1E3A5F')
    win  = PatternFill('solid', fgColor='D6F4DC')
    loss = PatternFill('solid', fgColor='FAD7D7')
    alt  = PatternFill('solid', fgColor='F2F2F2')
    brd  = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                  top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    cw   = {'Nr': 6, 'Datum': 18, 'Coin': 10, 'Symbol': 22, 'Timeframe': 12, 'Richtung': 10,
             'Ergebnis': 14, 'Entry-Preis': 14, 'Exit-Preis': 14, 'PnL (USDT)': 14, 'Gesamtkapital': 16}
    hdrs = list(rows[0].keys())
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = brd
        ws.column_dimensions[get_column_letter(c)].width = cw.get(h, 14)
    ws.row_dimensions[1].height = 22
    for ri, row in enumerate(rows, 2):
        # Verlierer sind immer rot -- vorher faelschlich nur auf geraden Zeilen (ri % 2 == 0),
        # ungerade Verlust-Zeilen bekamen still das neutrale Streifen-Grau statt Rot.
        f = win if row['Ergebnis'] == 'TP erreicht' else loss
        for c, key in enumerate(hdrs, 1):
            cell = ws.cell(row=ri, column=c, value=row[key])
            cell.fill = f
            cell.border = brd
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Entry-Preis', 'Exit-Preis'):
                cell.number_format = '#,##0.000000'
            elif key in ('PnL (USDT)', 'Gesamtkapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[ri].height = 18
    pnl = final.get('total_pnl_pct', 0)
    dd  = final.get('max_drawdown_pct', 0)
    wr  = final.get('win_rate', 0)
    eq  = final.get('end_capital', equity)
    n   = final.get('trade_count', len(trades))
    sr  = len(rows) + 3
    for label, val in [('Zeitraum', f'{start_date} -> {end_date}'), ('Trades', n),
                        ('Win-Rate', f'{wr:.1f}%'), ('PnL', f'{pnl:+.1f}%'),
                        ('Endkapital', f'{eq:.2f} USDT'), ('Max Drawdown', f'{dd:.1f}%')]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=val)
        sr += 1
    outfile = f'/tmp/{BOT_NAME}_trades.xlsx'
    wb.save(outfile)
    print(f'  {G}Excel erstellt: {outfile}{NC}')
    return outfile


# Palette + Marker-Optik von dnabot (run_portfolio_optimizer_momentum_exit.py) uebernommen,
# damit alle eigenen Bots visuell konsistente Reports/Charts liefern.
PAIR_COLORS = [
    '#f59e0b', '#8b5cf6', '#ec4899', '#14b8a6',
    '#f97316', '#84cc16', '#06b6d4', '#a78bfa',
]


def generate_equity_html(final, capital, start_date, end_date, labels):
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print(f'  {Y}plotly nicht installiert — Chart uebersprungen.{NC}')
        return None

    eq_df = final.get('equity_curve')
    if eq_df is None or (hasattr(eq_df, 'empty') and eq_df.empty):
        print(f'  {Y}Equity-Kurve leer — HTML-Chart uebersprungen.{NC}')
        return None

    # 'timestamp' kommt je nach Quelle als Spalte ODER als Index an: der
    # --auto-write-Pfad bezieht `final` aus src/ltbbot/analysis/portfolio_optimizer.py
    # (eigene interne Simulation, timestamp=Index), der --replot-Pfad aus
    # portfolio_simulator.py::run_portfolio_simulation (timestamp=Spalte).
    # reset_index() macht aus einem benannten Index wieder eine Spalte -- fuer
    # den Spalten-Fall ein No-Op (Index ist dann einfach 0..n, harmlos).
    if 'timestamp' not in eq_df.columns:
        eq_df = eq_df.reset_index()
    if 'timestamp' not in eq_df.columns:
        print(f'  {Y}Equity-Kurve ohne "timestamp" (Spalte/Index) — HTML-Chart uebersprungen.{NC}')
        return None

    times = [str(t) for t in eq_df['timestamp']]
    vals  = [float(v) for v in eq_df['equity']]
    pnl   = final.get('total_pnl_pct', 0)
    dd    = final.get('max_drawdown_pct', 0)
    wr    = final.get('win_rate', 0)
    n     = final.get('trade_count', 0)
    eq    = final.get('end_capital', vals[-1] if vals else capital)
    sign  = '+' if pnl >= 0 else ''
    title = (f"{BOT_NAME} Portfolio — {', '.join(labels)} | "
             f"PnL: {sign}{pnl:.1f}% | Equity: {eq:.2f} USDT | "
             f"MaxDD: {dd:.1f}% | WR: {wr:.1f}% | {n} Trades")

    fig = make_subplots(specs=[[{"secondary_y": True}]])

    # Einzel-Equity je Strategie (duenne, halbtransparente Linien) -- gleiche
    # Optik wie dnabot: zeigt auf einen Blick, welche Strategie das Portfolio
    # traegt, ohne die dicke Portfolio-Linie zu ueberdecken.
    trades_df = final.get('trades_df')
    if trades_df is not None and hasattr(trades_df, 'empty') and not trades_df.empty:
        pairs = sorted(set(zip(trades_df['symbol'], trades_df['timeframe'])))
        for idx, (sym, tf) in enumerate(pairs):
            sub = trades_df[(trades_df['symbol'] == sym) & (trades_df['timeframe'] == tf)] \
                    .sort_values('exit_time')
            peq    = capital
            ptimes = [str(sub.iloc[0]['entry_time'])] if len(sub) else []
            pvals  = [peq]
            for _, t in sub.iterrows():
                peq += float(t['pnl_usd'])
                ptimes.append(str(t['exit_time']))
                pvals.append(round(peq, 2))
            fig.add_trace(go.Scatter(
                x=ptimes, y=pvals, mode='lines', name=f"{sym.split('/')[0]}/{tf}",
                line=dict(color=PAIR_COLORS[idx % len(PAIR_COLORS)], width=1), opacity=0.55,
            ), secondary_y=False)

        # Entry-/Exit-Marker auf der Portfolio-Equity-Kurve
        entry_x, entry_y, entry_txt = [], [], []
        exit_win_x, exit_win_y = [], []
        exit_loss_x, exit_loss_y = [], []
        eq_running = capital
        for _, t in trades_df.sort_values('exit_time').iterrows():
            eq_running += float(t['pnl_usd'])
            tip = f"{t['symbol']} {t['timeframe']}<br>Equity: {eq_running:.2f} USDT"
            entry_x.append(str(t['entry_time'])); entry_y.append(eq_running); entry_txt.append(tip)
            if str(t.get('reason', '')).upper() == 'WIN':
                exit_win_x.append(str(t['exit_time'])); exit_win_y.append(eq_running)
            else:
                exit_loss_x.append(str(t['exit_time'])); exit_loss_y.append(eq_running)

        if entry_x:
            fig.add_trace(go.Scatter(
                x=entry_x, y=entry_y, mode='markers',
                marker=dict(color='#16a34a', symbol='triangle-up', size=12,
                            line=dict(width=1, color='#0f5132')),
                name='Entry ▲', text=entry_txt, hovertemplate='%{text}<extra>Entry</extra>',
            ), secondary_y=True)
        if exit_win_x:
            fig.add_trace(go.Scatter(
                x=exit_win_x, y=exit_win_y, mode='markers',
                marker=dict(color='#22d3ee', symbol='circle', size=10,
                            line=dict(width=1, color='#0e7490')),
                name='Exit TP ✓',
            ), secondary_y=True)
        if exit_loss_x:
            fig.add_trace(go.Scatter(
                x=exit_loss_x, y=exit_loss_y, mode='markers',
                marker=dict(color='#ef4444', symbol='x', size=10,
                            line=dict(width=2, color='#7f1d1d')),
                name='Exit SL ✗',
            ), secondary_y=True)

    fig.add_trace(go.Scatter(x=times, y=vals, mode='lines', name='Portfolio Equity',
                             line=dict(color='#2563eb', width=2), opacity=0.85),
                  secondary_y=True)
    fig.add_hline(y=capital, line=dict(color='rgba(100,100,100,0.4)', width=1, dash='dash'),
                  annotation_text=f'Start {capital:.0f} USDT', annotation_position='top left')
    fig.update_layout(title=dict(text=title, font=dict(size=12), x=0.5, xanchor='center'),
                      height=700, template='plotly_white', hovermode='x unified',
                      dragmode='zoom',
                      xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
                      legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5))
    fig.update_yaxes(title_text='Einzel-Equity (USDT)', secondary_y=False, fixedrange=False)
    fig.update_yaxes(title_text='Portfolio-Equity (USDT)', secondary_y=True, fixedrange=False)
    outfile = f'/tmp/{BOT_NAME}_portfolio_equity.html'
    fig.write_html(outfile)
    print(f'  {G}Chart erstellt: {outfile}{NC}')
    return outfile


def _do_replot(settings: dict, capital: float, start_date: str, end_date: str) -> int:
    print(f"\n{'─'*72}")
    print(f"{B}  ltbbot — Replot (aktives Portfolio){NC}")
    print(f"  Kapital: {capital:.0f} USDT | Zeitraum: {start_date} -> {end_date}")
    print(f"{'─'*72}\n")

    active = [s for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
              if s.get('active')]
    if not active:
        print(f"{R}  Keine aktiven Strategien in settings.json.{NC}")
        return 1

    active_pairs = {(s['symbol'], s['timeframe']) for s in active}
    matching = []
    for path in _scan_configs():
        try:
            with open(path) as f:
                cfg = json.load(f)
            m = cfg.get('market', {})
            if (m.get('symbol'), m.get('timeframe')) in active_pairs:
                matching.append(path)
        except Exception:
            pass

    if not matching:
        print(f"{R}  Keine Config-Dateien fuer aktive Strategien gefunden.{NC}")
        return 1

    strategies_data = _build_strategies_data(matching, start_date, end_date)
    if not strategies_data:
        print(f"{R}  Keine Daten geladen.{NC}")
        return 1

    from ltbbot.analysis.portfolio_simulator import run_portfolio_simulation
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        final = run_portfolio_simulation(capital, strategies_data, start_date, end_date,
                                          multi_band_entries=True)
    if not final:
        print(f"{R}  Portfolio-Simulation fehlgeschlagen.{NC}")
        return 1

    labels = [f"{sd.get('symbol', '?')}/{sd.get('timeframe', '?')}"
              for sd in strategies_data.values()]
    pnl = final.get('total_pnl_pct', 0)
    dd  = final.get('max_drawdown_pct', 0)
    n   = final.get('trade_count', 0)
    wr  = final.get('win_rate', 0)
    eq  = final.get('end_capital', 0)

    print(f"\n{'='*72}")
    print(f"{B}  Replot — {len(matching)} Strategie(n){NC}\n")
    for fname, sd in strategies_data.items():
        print(f"  {G}OK{NC} {sd.get('symbol', fname):<26} / {sd.get('timeframe', ''):<6}")
    print(f"\n  Endkapital: {eq:.2f} USDT  | PnL: {pnl:+.1f}%  | MaxDD: {dd:.2f}%")
    print(f"{'='*72}\n")

    summary = (f"{BOT_NAME} Replot\n"
               f"{len(matching)} Strategien | {n} Trades | WR: {wr:.1f}%\n"
               f"PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}% | Equity: {eq:.2f} USDT\n"
               f"Zeitraum: {start_date} -> {end_date}")
    _send_telegram(summary)
    xlsx = generate_trades_excel(final, strategies_data, capital, start_date, end_date)
    if xlsx:
        _send_telegram_doc(xlsx, caption=f'{BOT_NAME} Trades | {n} Trades | WR: {wr:.1f}% | Equity: {eq:.2f} USDT')
    html = generate_equity_html(final, capital, start_date, end_date, labels)
    if html:
        _send_telegram_doc(html, caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}%')
    return 0


def compute_min_capital(portfolio_files: list, strategies_data: dict, safety_margin_pct: float = 20.0):
    """
    Berechnet je gewaehlter Strategie und Envelope-Band das Mindestkapital,
    damit die Positionsgroesse (risk_amount_usd / SL-Abstand) Bitgets
    5-USDT-Notional-Minimum erreicht (User-Anforderung 2026-08-27: "nach
    allen Regeln" -- alle 3 Baender jeder gewaehlten Strategie muessen
    einzeln die Mindestgroesse erreichen koennen, da bei multi_band_entries
    jedes Band unabhaengig eine eigene Position eroeffnen kann).

    Formel (sl_mode='ratio', Standard seit dem Optuna-Port in optimizer.py):
      risk_amount_usd = capital * risk_per_entry_pct/100
      notional         = risk_amount_usd / sl_pct  (sl_pct = SL-Abstand vom Entry)
      notional >= 5  =>  capital >= 500 * sl_pct / risk_per_entry_pct
    Nur der GROESSTE Wert ueber alle (Strategie, Band)-Kombinationen zaehlt --
    das gemeinsame Kapital muss fuer die anspruchsvollste Kombination reichen,
    nicht im Schnitt. ATR-basierte SL-Configs (marktabhaengig, keine feste
    Formel ohne aktuelle Kursdaten) werden separat markiert, nicht berechnet.

    Rueckgabe: (min_capital_exact, min_capital_recommended_mit_puffer, details-Liste)
    """
    MIN_NOTIONAL_USDT = 5.0
    details = []
    min_capital_exact = 0.0
    for fname in portfolio_files:
        sd = strategies_data.get(fname, {})
        cfg = sd.get('params', {}) or {}
        strat = cfg.get('strategy', {})
        risk = cfg.get('risk', {})
        symbol = sd.get('symbol', fname)
        timeframe = sd.get('timeframe', '?')
        risk_pct = risk.get('risk_per_entry_pct', 0.5)
        envelopes = strat.get('envelopes', [])
        if 'sl_to_env1_ratio' in risk and envelopes and risk_pct > 0:
            sl_ratio = risk['sl_to_env1_ratio']
            worst_band, worst_cap = None, 0.0
            for k, env_pct in enumerate(envelopes, 1):
                sl_pct = env_pct * sl_ratio
                if sl_pct <= 0:
                    continue
                cap_needed = MIN_NOTIONAL_USDT * 100.0 * sl_pct / risk_pct
                if cap_needed > worst_cap:
                    worst_cap, worst_band = cap_needed, k
            details.append({'symbol': symbol, 'timeframe': timeframe,
                             'worst_band': worst_band, 'min_capital': round(worst_cap, 2)})
            min_capital_exact = max(min_capital_exact, worst_cap)
        else:
            details.append({'symbol': symbol, 'timeframe': timeframe,
                             'worst_band': None, 'min_capital': None,
                             'note': 'ATR-basierter SL -- marktabhaengig, nicht analytisch berechenbar'})

    min_capital_recommended = min_capital_exact * (1 + safety_margin_pct / 100.0) if min_capital_exact > 0 else 0.0
    return min_capital_exact, min_capital_recommended, details


def main() -> int:
    parser = argparse.ArgumentParser(description='ltbbot Portfolio-Optimizer (Envelope)')
    parser.add_argument('--capital',    type=float, default=None)
    parser.add_argument('--max-dd',     type=float, default=30.0)
    parser.add_argument('--start-date', type=str,   default=None)
    parser.add_argument('--end-date',   type=str,   default=None)
    parser.add_argument('--auto-write', action='store_true')
    parser.add_argument('--replot',     action='store_true',
                        help='Replot fuer aktives Portfolio (keine Re-Optimierung)')
    args = parser.parse_args()

    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    opt       = settings.get('optimization_settings', {})
    capital   = args.capital or float(opt.get('start_capital', 50))
    max_dd    = args.max_dd

    # Lookback (in Tagen) IMMER zuerst bestimmen -- backtest_lookback_weeks hat
    # Vorrang (aus Walk-Forward-Analyse), sonst LOOKBACK_MAP-Fallback ueber die
    # aktiven Timeframes. Dieser EINE Wert treibt sowohl start_date als auch die
    # OOS-Reserve unten -- vorher wurden beide unabhaengig voneinander berechnet
    # (Lookback ueber backtest_lookback_weeks, OOS-Reserve ueber das theoretische
    # Timeframe-Maximum aus LOOKBACK_MAP), wodurch bei kurzem backtest_lookback_weeks
    # (z.B. 1 Woche) end_date (Trainingsende, OOS-Reserve-abhaengig) weiter in der
    # Vergangenheit lag als start_date -- ein rueckwaerts laufendes, leeres Fenster.
    lookback_weeks = opt.get('backtest_lookback_weeks')
    if lookback_weeks:
        lookback = int(lookback_weeks) * 7
        print(f"  Lookback: {lookback_weeks} Wochen (aus backtest_lookback_weeks)")
    else:
        active_tfs = [
            s.get('timeframe', '1h')
            for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
            if s.get('active', True)
        ]
        lookback = max((LOOKBACK_MAP.get(tf, 365) for tf in active_tfs), default=365)

    start_date = args.start_date or (date.today() - timedelta(days=lookback)).strftime('%Y-%m-%d')

    # OOS: 70/30-Split, Portfolio-Optimizer sieht nur Training. Referenzdatum ist
    # IMMER "heute" (rollierend) -- User-Entscheidung 2026-08-27: das statische
    # `oos_reference_date` aus settings.json (eingefuehrt 2026-06-21, Commit 6174880,
    # Zweck: "Portfolio-Optimizer sieht niemals OOS-Daten") wurde nie aktualisiert
    # und fror das OOS-Fenster dauerhaft auf den Einfuehrungs-Zeitpunkt ein. Die
    # OOS-Reserve ist jetzt 30% des TATSAECHLICH genutzten Lookback-Fensters
    # (nicht mehr des theoretischen Timeframe-Maximums) -- bleibt dadurch immer
    # konsistent mit start_date, egal wie lang/kurz backtest_lookback_weeks ist.
    if opt.get('oos_reference_date') and not args.end_date:  # Vorhandensein des Keys = OOS-Schutz aktiv/deaktivierbar
        ref_dt       = date.today()
        oos_days     = max(1, lookback * 30 // 100)
        oos_start_dt = ref_dt - timedelta(days=oos_days)
        end_date     = (oos_start_dt - timedelta(days=1)).strftime('%Y-%m-%d')
        print(f"  OOS 70/30: ref={ref_dt} (rollierend=heute) oos_start={oos_start_dt} end={end_date}")
    else:
        end_date = args.end_date or date.today().strftime('%Y-%m-%d')

    if args.replot:
        return _do_replot(settings, capital, start_date, end_date)

    print(f"\n{'─'*72}")
    print(f"{B}  ltbbot — Automatische Portfolio-Optimierung (Envelope){NC}")
    print(f"  Greedy-Selektion | MaxDD <= {max_dd:.0f}% | Kapital: {capital:.0f} USDT")
    print(f"  Zeitraum: {start_date} -> {end_date}")
    print(f"{'─'*72}\n")

    config_files = _scan_configs()
    if not config_files:
        print(f"{R}  Keine *_envelope.json Configs in {CONFIGS_DIR}{NC}")
        print(f"  -> Zuerst run_pipeline.sh ausfuehren!\n")
        return 1

    print(f"  {len(config_files)} Config(s) gefunden.\n")
    strategies_data = _build_strategies_data(config_files, start_date, end_date)
    if not strategies_data:
        print(f"{R}  Keine Daten geladen.{NC}")
        return 1

    # Rangfolge-Glaettung: Wahl des "Star-Spielers" basiert auf dem Mittel mehrerer
    # versetzter Trailing-Snapshots statt nur dem Stichtag (validiert in zerobot:
    # Calmar 20.7 vs. 14.1 Baseline, OOS-Test).
    smoothing_step_days = int(opt.get('smoothing_step_days', 2))
    smoothing_samples   = int(opt.get('smoothing_samples', 7))

    from ltbbot.analysis.portfolio_optimizer import run_portfolio_optimizer
    result = run_portfolio_optimizer(capital, strategies_data, start_date, end_date,
                                      max_portfolio_dd_constraint=max_dd / 100.0,
                                      smoothing_step_days=smoothing_step_days,
                                      smoothing_samples=smoothing_samples)

    if not result or not result.get('optimal_portfolio'):
        print(f"{R}  Kein Portfolio erfuellt die Bedingungen (MaxDD <= {max_dd:.0f}%).{NC}\n")
        return 0

    portfolio_files = result['optimal_portfolio']
    final           = result.get('final_result') or {}

    print(f"\n{'='*72}")
    print(f"{B}  Optimales Portfolio — {len(portfolio_files)} Strategie(n){NC}\n")
    for fname in portfolio_files:
        sd = strategies_data.get(fname, {})
        print(f"  {G}OK{NC} {sd.get('symbol', fname):<26} / {sd.get('timeframe', ''):<6}")
    if final:
        pnl = final.get('total_pnl_pct', 0)
        print(f"\n  Endkapital: {final.get('end_capital', 0):.2f} USDT  "
              f"| PnL: {pnl:+.1f}%  "
              f"| MaxDD: {final.get('max_drawdown_pct', 0):.2f}%")

    min_cap_exact, min_cap_reco, min_cap_details = compute_min_capital(portfolio_files, strategies_data)
    if min_cap_exact > 0:
        print(f"\n  {B}Mindestkapital (alle Baender jeder Strategie erreichen 5-USDT-Notional):{NC}")
        for d in min_cap_details:
            if d.get('min_capital') is not None:
                print(f"    {d['symbol']:<20} / {d['timeframe']:<4}  Band {d['worst_band']} massgeblich  "
                      f"->  {d['min_capital']:.2f} USDT")
            else:
                print(f"    {d['symbol']:<20} / {d['timeframe']:<4}  {d.get('note', '')}")
        print(f"  Exaktes Minimum: {min_cap_exact:.2f} USDT  |  Empfohlen (+20% Puffer): {min_cap_reco:.2f} USDT")
    print(f"{'='*72}\n")

    # Vergleich mit aktuellem Portfolio
    cur_result = _simulate_current_portfolio(settings, strategies_data, capital, start_date, end_date)
    cur_cap    = cur_result.get('end_capital', 0) if cur_result else 0
    new_cap    = final.get('end_capital', 0)
    if cur_result:
        print(f"  Aktuelles Portfolio: {cur_cap:.2f} USDT  "
              f"| PnL: {cur_result.get('total_pnl_pct', 0):+.1f}%  "
              f"| MaxDD: {cur_result.get('max_drawdown_pct', 0):.2f}%")
        print(f"  Neues Portfolio:     {new_cap:.2f} USDT  "
              f"| PnL: {final.get('total_pnl_pct', 0):+.1f}%  "
              f"| MaxDD: {final.get('max_drawdown_pct', 0):.2f}%\n")

    if args.auto_write:
        if cur_result and new_cap <= cur_cap:
            print(f"{Y}  Neues Portfolio ({new_cap:.2f} USDT) nicht besser als aktuelles "
                  f"({cur_cap:.2f} USDT) — keine Aenderung.{NC}\n")
        else:
            _write_to_settings(portfolio_files, strategies_data)
            print(f"{G}  settings.json aktualisiert — {len(portfolio_files)} Strategie(n).{NC}\n")
    else:
        current_set = {
            (s.get('symbol'), s.get('timeframe'))
            for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
            if s.get('active')
        }
        new_set = {
            (strategies_data.get(f, {}).get('symbol'), strategies_data.get(f, {}).get('timeframe'))
            for f in portfolio_files
        }
        if current_set == new_set:
            print(f"{Y}  Portfolio unveraendert — keine Aenderung noetig.{NC}\n")
        else:
            try:
                ans = input("  Optimales Portfolio in settings.json eintragen? (j/n): ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                ans = 'n'
            if ans in ('j', 'ja', 'y', 'yes'):
                _write_to_settings(portfolio_files, strategies_data)
                print(f"{G}  settings.json aktualisiert.{NC}\n")
            else:
                print(f"{Y}  settings.json NICHT geaendert.{NC}\n")

    # Reports + Telegram (nur im auto-write Modus)
    if args.auto_write:
        labels = [
            f"{strategies_data.get(f, {}).get('symbol', '?')}/{strategies_data.get(f, {}).get('timeframe', '?')}"
            for f in portfolio_files
        ]
        pnl = final.get('total_pnl_pct', 0)
        dd  = final.get('max_drawdown_pct', 0)
        n   = final.get('trade_count', 0)
        wr  = final.get('win_rate', 0)
        eq  = final.get('end_capital', 0)
        summary = (f"{BOT_NAME} Auto-Optimizer\n"
                   f"{len(portfolio_files)} Strategien | {n} Trades | WR: {wr:.1f}%\n"
                   f"PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}% | Equity: {eq:.2f} USDT\n"
                   f"Zeitraum: {start_date} -> {end_date}")
        if min_cap_exact > 0:
            summary += f"\n\nMindestkapital: {min_cap_exact:.2f} USDT (empfohlen mit Puffer: {min_cap_reco:.2f} USDT)"
        _send_telegram(summary)
        xlsx = generate_trades_excel(final, strategies_data, capital, start_date, end_date)
        if xlsx:
            _send_telegram_doc(xlsx, caption=f'{BOT_NAME} Trades | {n} Trades | WR: {wr:.1f}% | Equity: {eq:.2f} USDT')
        html = generate_equity_html(final, capital, start_date, end_date, labels)
        if html:
            _send_telegram_doc(html, caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}%')

    return 0


if __name__ == '__main__':
    sys.exit(main())
